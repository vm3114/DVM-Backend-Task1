import csv
import openpyxl

def have_common_elements(list1, list2):
    set1 = set(list1)
    set2 = set(list2)
    
    if (set1 & set2):
        return True
    else:
        return False


class Course:

    def _init_(self, name, code, test_date, credits):
        self.name = name
        self.code = code
        self.credits = credits
        self.test_date = test_date
        self.sections = []

    def get_all_sections(self):
        return self.sections

    def _str_(self):
        pass

    def populate_section(self, section):
        self.sections.append(section)
    

class Section:

    def _init_(self, course, id, room, day, hours):
        self.course = course
        self.section_id = id
        self.room = room
        self.day = day
        self.hours = hours
        if id[0] == "L":
            self.type = "Lecture"
        elif id[0] == "T":
            self.type = "Tutorial"
        elif id[0] == "P":
            self.type = "Practise"



class Timetable:

    def _init_(self):
        self.courses = []
        self.enrolled_sections = []
    
    def enroll_subject(self, course):
        self.courses.append(course)
    
    def check_clashes(self, section):
        
        for course in self.courses:
            if section in course.get_all_sections():
                for enrolled_sections in self.enrolled_sections:
                    if enrolled_sections.course != section.course:
                            if have_common_elements(enrolled_sections.day, section.day):
                                if have_common_elements(enrolled_sections.hours, section.hours):
                                    return True
                                else:
                                    return False
                            else:
                                return False
                    else:
                        if enrolled_sections.type != section.type:
                            if have_common_elements(enrolled_sections.day, section.day):
                                if have_common_elements(enrolled_sections.hours, section.hours):
                                    return True
                                else:
                                    return False
                            else:
                                return False
        return True
                            
    def export_to_csv(self, filename):
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Course', 'Section Number', 'Room Number', 'Examination Date', 'Day', 'Slot', 'Type'])

            for course in self.courses:
                for section in course.get_all_sections():
                    writer.writerow([course.name, section.section_id, section.room, course.test_date, section.day, section.hours, section.type])
        
def extract_timetable(info):
    with open(info, 'r') as i:
        reader = csv.reader(i)
        rowlist = list(reader)
        rowlist.pop(0)
        return(rowlist)

timetable = Timetable()

def populate_course(file_path, timetable):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row = 2, values_only = True):
        course_id, name, test_date, course_credits = row
        ncourse = Course(name, course_id, test_date, course_credits)
        timetable.enroll_subject(ncourse)
    workbook.save(file_path)

def extract_and_populate_sections(file_path, course):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row = 2, values_only = True):
        course_name, course_id, section_id, room_no, day_space, hour_space = row
        days = day_space.split(" ")
        hours = hour_space.split(" ")
        nsection = Section(course_name, section_id, room_no, days, hours)
        course.populate_course(nsection)
    workbook.save(file_path)
    
